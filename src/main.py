from bot.bot import Bot
import openpyxl
import os

# Create bot and load webpage
website = "https://rpachallenge.com/"
print(f"Opening website: {website}")
bot = Bot()
bot.open_new_page(website)

spreadsheet_location = "/data/challenge.xlsx"
# Will need to append a . for non-alpine contexts
if not os.getenv("DOCKERIZED"):
  spreadsheet_location = '.' + spreadsheet_location

# Load spreadsheet of data to be entered into RPA challenge site
print("Loading spreadsheet")
workbook = openpyxl.load_workbook(spreadsheet_location)
sheet = workbook.active

# Click start button to begin challenge
print("Starting task. . .")
start_button = bot.find_xpath("//button[text()='Start']")
start_button.click()

def get_submit():
  return bot.find_xpath("//input[@value='Submit']")

submit_button = get_submit()

# The row from which we'll get the data to insert into the specified field
input_data_row = 2

# As long as the submit button is visible
while submit_button:

  # Loop through the 7 columns of the table, one for each input field
  for i in range(1,8):
    # Get the current column header
    target_text = str(sheet.cell(row=1, column=i).value).strip()

    # Get the input that has the same text as the column header
    target_input = bot.find_xpath(f"//label[text()='{target_text}']/following-sibling::input")

    # Read the text to insert from the table
    input_text = sheet.cell(row=input_data_row, column=i).value

    # Put the input text into the form field
    target_input.send_keys(input_text)

  # Click the button and then try to see if the button still exists
  submit_button.click()

  # Try to find the submit button. If it isn't present, the task is completed so end the loop
  try:
    submit_button = get_submit()
  except:
    break

  input_data_row += 1

# Get the result message
print("Task completed")
result_element = bot.find_class("message2")
result = result_element.get_attribute("innerHTML")
print(f"Results: {result}")

# Shut down the webpage
print("Shutting down. . .")
bot.close_page()


''' 

Simple form test site:
https://formy-project.herokuapp.com/

Advanced form test site:
https://rpachallenge.com/

'''